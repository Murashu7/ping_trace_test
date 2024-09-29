import subprocess
import re
import pandas as pd
import datetime
import os
import asyncio
from aiofiles import open as aio_open  # 非同期でファイルを開く
import yaml
import openpyxl
from openpyxl.styles import PatternFill
import platform
import locale

def load_config(config_path):
    """
    設定ファイルを読み込む関数
    
    Parameters:
    - config_path (str): 設定ファイルのパス
    
    Returns:
    - config (dict): 設定内容を含む辞書
    """
    with open(config_path, 'r') as file:
        config = yaml.safe_load(file)
    
    ping_count = config.get('ping_count', 5) # デフォルト値
    average_rtt = config.get('average_rtt', 100)  # デフォルト値
    max_packet_loss = config.get('max_packet_loss', 20)  # デフォルト値
    max_hopu = config.get('max_hopu', 10)  # デフォルト値
    
    return ping_count, average_rtt, max_packet_loss, max_hopu

# 非同期で結果をログファイルに保存する関数
async def save_results(kyoten_name, test_type, index, host, output, results_dir, result_type):
    """pingまたはtracerouteの結果をログファイルに保存"""
    full_output_path = os.path.join(results_dir, f'{index}_{kyoten_name}_{test_type}の結果.log')
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 結果があればログファイルに書き込む
    if output:
        async with aio_open(full_output_path, 'a') as f:
            await f.write(f"{host} への {result_type} 結果 {current_time}\n\n")
            await f.write(f"{output}")
            await f.write("\n" + "="*60 + "\n")

# async def ping(count_opt, ping_count, host):
#     """pingコマンドを非同期で実行"""
#     proc = await asyncio.create_subprocess_exec(
#         'ping', count_opt, str(ping_count), host, 
#         stdout=subprocess.PIPE, 
#         stderr=subprocess.PIPE
#     )
#     stdout, stderr = await proc.communicate()
#     return stdout.decode() if stdout else stderr.decode()

async def ping(cmd_args):
    """pingコマンドを非同期で実行"""
    proc = await asyncio.create_subprocess_exec(
        *cmd_args, 
        stdout=subprocess.PIPE, 
        stderr=subprocess.PIPE
    )
    stdout, stderr = await proc.communicate()
    return stdout.decode() if stdout else stderr.decode()

def evaluate_ping_result(packet_loss, avg_rtt, average_rtt, max_packet_loss):
    """
    パケットロス率とRTTの結果を判定
    """
    if packet_loss <= max_packet_loss and avg_rtt <= average_rtt:
        print(f"成功 (パケットロス: {packet_loss}%, 平均RTT: {avg_rtt} ms)")
        return True
    else:
        print(f"失敗 (パケットロス: {packet_loss}%, 平均RTT: {avg_rtt} ms)")
        return False

def parse_packet_loss(output, pattern):
    """
    パケットロス率の解析
    Parameters:
    - output (str): pingの結果出力
    - pattern (str): パケットロス率の正規表現パターン
    Returns:
    - packet_loss (int): パケットロス率、解析失敗時はNone
    """
    packet_loss_match = re.search(pattern, output)
    if packet_loss_match:
        return float(packet_loss_match.group(1))  # パケットロス率を取得
    else:
        print("パケットロス率の解析に失敗しました。")
        return None

def parse_rtt(output, pattern):
    """
    RTTの解析
    Parameters:
    - output (str): pingの結果出力
    - pattern (str): RTTの正規表現パターン
    Returns:
    - avg_rtt (int): 平均RTT、解析失敗時はNone
    """
    rtt_match = re.search(pattern, output)
    if rtt_match:
        return float(rtt_match.group(1))  # 平均RTTを取得
    else:
        print("RTTの解析に失敗しました。")
        return None

# TODO: Windowsでのping判定
# def validate_ping_windows(expected_status, is_ja, output, average_rtt, max_packet_loss):
#     if is_ja: # 日本語版WindowsのパケットロスとRTT解析
#         if expected_status == 'ng':
#             if "要求がタイムアウトしました" in output or "宛先ホストに到達できません" in output:
#                 return True
#             else:
#                 return False
#         else:
#             packet_loss = parse_packet_loss(output, r'(\d+)% の損失')
#             avg_rtt = parse_rtt(output, r'平均 = (\d+)ms')
        
#     else: # 英語版WindowsのパケットロスとRTT解析
#         if expected_status == 'ng':
#             if "Request timed out" in output or "Destination host unreachable" in output:
#                 return True
#             else:
#                 return False
#         else:
#             packet_loss = parse_packet_loss(output, r'(\d+)% loss')
#             avg_rtt = parse_rtt(output, r'Average = (\d+)ms')
    
#     if packet_loss is None or avg_rtt is None:
#         return False
    
#     # パケットロスとRTTの判定
#     return evaluate_ping_result(packet_loss, avg_rtt, average_rtt, max_packet_loss)

# TODO: Windowsでのping判定
def validate_ping_windows(expected_status, is_ja, output, average_rtt, max_packet_loss):
    error_msgs = {
        'ja': ["要求がタイムアウトしました", "宛先ホストに到達できません"],
        'en': ["Request timed out", "Destination host unreachable"]
    }

    lang = 'ja' if is_ja else 'en'

    if expected_status == 'ng':
        return any(msg in output for msg in error_msgs[lang])

    packet_loss_pattern = r'(\d+)% の損失' if is_ja else r'(\d+)% loss'
    avg_rtt_pattern = r'平均 = (\d+)ms' if is_ja else r'Average = (\d+)ms'
    
    packet_loss = parse_packet_loss(output, packet_loss_pattern)
    avg_rtt = parse_rtt(output, avg_rtt_pattern)

    if packet_loss is None or avg_rtt is None:
        return False

    return evaluate_ping_result(packet_loss, avg_rtt, average_rtt, max_packet_loss)

# Macでのping判定（言語によってメッセージは変わらない）
def validate_ping_mac(expected_status, output, average_rtt, max_packet_loss):
    # TODO: expected_status が ng かつ ping が通った場合は False とする
    if expected_status == 'ng':
        if "Request timeout" in output or "Destination Host Unreachable" in output:
            return True
        else:
            return False
            
    packet_loss = parse_packet_loss(output, r'(\d+)\.\d+% packet loss')
    avg_rtt = parse_rtt(output, r'round-trip min/avg/max/stddev = [\d.]+/([\d.]+)/[\d.]+/[\d.]+ ms')
    
    if packet_loss is None or avg_rtt is None:
        return False
     
    # パケットロスとRTTの判定
    return evaluate_ping_result(packet_loss, avg_rtt, average_rtt, max_packet_loss)
      
# pingの出力と期待される結果を比較
def validate_ping(is_win, is_ja, average_rtt, max_packet_loss, output, expected_status):    
    if output is None:
        return False
        
    # WindowsとMacの解析処理を分岐
    if is_win:
        return validate_ping_windows(expected_status, is_ja, output, average_rtt, max_packet_loss)
    else:
        return validate_ping_mac(expected_status, output, average_rtt, max_packet_loss)

# 非同期でホストへpingを送信し、結果を評価する関数
async def ping_and_validate(is_win, is_ja, ping_count, average_rtt, max_packet_loss, host, expected_status):
    cmd = 'ping'
    opt = '-n' if is_win else '-c'
    cmd_args = [cmd, opt, str(ping_count), host]
        
    """ホストにpingを送信して、結果を評価"""
    print(f"{host} へのpingを確認中...")
    ping_output = await ping(cmd_args)  # ping関数は非同期で実行
    print(ping_output)

    # pingの結果をバリデート
    success = validate_ping(is_win, is_ja, average_rtt, max_packet_loss, ping_output, expected_status)
    
    if success:
        print(f"{host} へのpingの判定は成功しました\n")
    else:
        print(f"{host} へのpingの判定は失敗しました\n")
    
    # pingの結果と評価結果を返す
    return ping_output, success

async def process_ping(is_win, is_ja, ping_count, average_rtt, max_packet_loss, kyoten_name, test_type, index, host, expected_status, results_dir):
    """個別のホストのpingを処理"""
    # pingを送信して結果を評価
    ping_output, success = await ping_and_validate(is_win, is_ja,ping_count, average_rtt, max_packet_loss, host, expected_status)

    # 結果をログファイルに保存
    await save_results(kyoten_name, test_type, index, host, ping_output, results_dir, 'ping')

    # 成功/失敗の結果を返す
    return {host: success}

async def ping_multiple_hosts(is_win, is_ja, ping_count, average_rtt, max_packet_loss, kyoten_name, test_type, list_ping_eval, results_dir):
    tasks = []
    for index, dict_ping_eval in enumerate(list_ping_eval, start=1):
        host = list(dict_ping_eval.keys())[0]
        expected_status = dict_ping_eval[host]
        task = asyncio.create_task(process_ping(is_win, is_ja, ping_count, average_rtt, max_packet_loss,kyoten_name, test_type, index, host, expected_status, results_dir))
        tasks.append(task)
    return await asyncio.gather(*tasks)

# IPアドレスを抽出する正規表現
ip_pattern = re.compile(r'(\d+\.\d+\.\d+\.\d+)')

# IPアドレスを抽出して配列に格納
def extract_ips(trace_result):
    ip_addresses = []
    for hop in trace_result:
        ip_match = ip_pattern.search(hop)
        if ip_match:
            ip_addresses.append(ip_match.group(0))
    return ip_addresses

# 共通の要素があるかどうかをチェックする関数
def check_route_match(list_ip, list_expected_route):
    # リストをsetに変換して共通要素を探す
    return bool(set(list_ip) & set(list_expected_route))

def validate_route(output, list_expected_route):
    """
    Tracerouteの出力と期待される経路を比較
    output: トレースルートの出力
    list_expected_route: 経路として期待されるホストリスト
    """
    if output is None:
        return False
    # 各ホップのIPアドレスを抽出
    list_ip = extract_ips(output.splitlines())
    return check_route_match(list_ip, list_expected_route)

async def traceroute(cmd_args):
    """tracerouteコマンドを非同期で実行"""
    proc = await asyncio.create_subprocess_exec(
        *cmd_args, 
        stdout=subprocess.PIPE, 
        stderr=subprocess.PIPE
    )
    stdout, _ = await proc.communicate()
    return stdout.decode()

# 非同期でホストへのtracerouteを実行し、結果を評価する関数
async def trace_and_validate(is_win, max_hop, host, list_expected_route):
    cmd = 'tracert' if is_win else 'traceroute'
    opt = f'-h {max_hop}' if is_win else ('-n', '-m', str(max_hop))
    cmd_args = [cmd, opt, host]if is_win else [cmd] + list(opt) + [host]

    """ホストにtracerouteを実行して、結果を評価"""
    print(f"{host} の経路を確認中...")
    trace_output = await traceroute(cmd_args)  # traceroute関数は非同期で実行
    print(trace_output)

    # 経路の結果をバリデート
    success = validate_route(trace_output, list_expected_route)
    
    if success:
        print(f"{host} は期待される経路を通過しました\n")
    else:
        print(f"{host} は期待される経路を通過しませんでした\n")
    
    # tracerouteの結果と評価結果を返す
    return trace_output, success

async def process_trace(is_win, max_hop, kyoten_name, test_type, index, host, list_expected_route, results_dir):
    """個別のホストのtracerouteを処理"""
    # tracerouteを実行して結果を評価
    trace_output, success = await trace_and_validate(is_win, max_hop, host, list_expected_route)

    # 結果をログファイルに保存
    await save_results(kyoten_name, test_type, index, host, trace_output, results_dir, "traceroute")

    # 成功/失敗の結果を返す
    return {host: success}

async def trace_multiple_hosts(is_win, max_hop, kyoten_name, test_type, list_trace_eval, results_dir):
    tasks = []
    for index, dict_trace_eval in enumerate(list_trace_eval, start=1):
        host = list(dict_trace_eval.keys())[0]
        list_expected_route = dict_trace_eval[host]
        task = asyncio.create_task(process_trace(is_win, max_hop, kyoten_name, test_type, index, host, list_expected_route, results_dir))
        tasks.append(task)
    
    return await asyncio.gather(*tasks)

async def ping_trace_multiple_hosts(is_win, is_ja, ping_count, average_rtt, max_packet_loss, max_hop, list_ping_eval, list_trace_eval, selected_kyoten_name, selected_test_type):
    ping_results = {}
    trace_results = {}
    
    # ファイルを一階層上の "results/selected_kyoten_name" フォルダに保存
    results_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', f'results/{selected_kyoten_name}/{selected_test_type}')
    
    # フォルダが存在しなければ作成
    os.makedirs(results_dir, exist_ok=True)
    
    ping_results = await ping_multiple_hosts(is_win, is_ja, ping_count, average_rtt, max_packet_loss, selected_kyoten_name, selected_test_type, list_ping_eval, results_dir)
    trace_results = await trace_multiple_hosts(is_win, max_hop, selected_kyoten_name, selected_test_type, list_trace_eval, results_dir)
    return ping_results, trace_results, results_dir

# 結果を辞書に変換する関数
def convert_to_list(df):
    list_ping_eval = []
    list_trace_eval = []
    
    for _, row in df.iterrows():
        dict_ping_eval = {}
        dict_trace_eval = {}
        dest = row['dest']
        dict_ping_eval[dest] = row['ping_eval']
        dict_trace_eval[dest] = [row['trace_eval_1'], row['trace_eval_2'], row['trace_eval_3']]
        list_ping_eval.append(dict_ping_eval)
        list_trace_eval.append(dict_trace_eval)
    return list_ping_eval, list_trace_eval

def select_kyoten(df):
    # 画面に全てのnumberとnameを表示
    print("全てのnumberとnameを表示します:")
    for index, row in df.iterrows():
        print(f"拠点番号: {row['number']}, {row['area']}, {row['name']}")

    # ユーザーが正しいnumberを選択するまでループ
    selected_row = None
    while selected_row is None or selected_row.empty:
        try:
            # numberを選択するメッセージを表示
            selected_number = int(input("試験をする拠点番号を選択してください: "))
            
            # 選択されたnumberに対応する行を取得
            selected_row = df[df['number'] == selected_number]
            
            if selected_row.empty:
                print("選択したnumberは存在しません。もう一度選択してください。")
        except ValueError:
            print("無効な入力です。数字を入力してください。")

    # 正しいnumberが選択されたらnameとtypeを変数に格納
    # selected_area = selected_row['area'].values[0]
    selected_kyoten_name = selected_row['name'].values[0]
    selected_kyoten_type = selected_row['type'].values[0]
    print(f"選択された拠点: {selected_kyoten_name}({selected_kyoten_type})")
    return selected_kyoten_name, selected_kyoten_type

def select_test_type(selected_kyoten_type, csv_file='../settings/test_info.csv'):
    # CSVファイルを読み込み
    try:
        df = pd.read_csv(csv_file)
    except FileNotFoundError:
        print("CSVファイルが見つかりません。")
        return None

    # 指定されたkyoten_typeに基づいて試験タイプを取得
    test_types = df[df['kyoten_type'] == selected_kyoten_type]['test_type'].tolist()
    if not test_types:
        print("選択された拠点には試験が設定されていません。")
        return None

    # 試験内容を表示
    print("試験内容を表示します:")
    for index, test_type in enumerate(test_types, start=1):
        print(f"番号: {index}, {test_type}")

    # ユーザーが正しいnumberを選択するまでループ
    selected_row = None
    while selected_row is None:
        try:
            # numberを選択するメッセージを表示
            selected_number = int(input("試験番号を選択してください: "))
            
            if 1 <= selected_number <= len(test_types):
                selected_row = test_types[selected_number - 1]
            else:
                print("選択したnumberは存在しません。もう一度選択してください。")
        except ValueError:
            print("無効な入力です。数字を入力してください。")

    print(f"選択された試験: {selected_row}")
    return selected_row

class SelectionError(Exception):
    """選択された拠点タイプまたは試験タイプが無効な場合に投げられる例外."""
    pass

def get_test_type_path(selected_kyoten_type, selected_test_type, test_info_path='../settings/test_info.csv'):
    """
    test_info.csvを読み込み、選択された拠点タイプと試験タイプに対応するCSVファイルのパスを取得する関数。
    
    Parameters:
    - selected_kyoten_type (str): 選択された拠点タイプ（例: "大規模", "中規模", "小規模"）
    - selected_test_type (str): 選択された試験タイプ
    - test_info_path (str): test_info.csvのパス（デフォルト: '../settings/test_info.csv'）
    
    Returns:
    - csv_path (str): 該当するテスト結果ファイルのパス
    """
    # test_info.csvの読み込み
    df_test_info = pd.read_csv(test_info_path)

    # selected_kyoten_type と selected_test_type にマッチする file_name を取得
    matching_row = df_test_info[
        (df_test_info['kyoten_type'] == selected_kyoten_type) &
        (df_test_info['test_type'] == selected_test_type)
    ]
    
    if not matching_row.empty:
        return f"../settings/{matching_row['file_name'].values[0]}"
    else:
        raise SelectionError("選択された拠点タイプまたは試験タイプが見つかりませんでした。")

def generate_unique_filename(results_dir, selected_kyoten_name, selected_test_type):
    """
    ファイル名を生成し、重複があれば連番を付与して一意なファイル名を生成する関数。
    
    Parameters:
    - results_dir (str): 保存するディレクトリパス
    - selected_kyoten_name (str): 拠点名
    - selected_test_type (str): 試験タイプ
    
    Returns:
    - excel_file (str): 一意なファイル名
    """
    # 基本となるファイル名
    base_filename = f'{selected_kyoten_name}_{selected_test_type}_判定表'
    extension = '.xlsx'
    
    # 最初のファイル名
    excel_file = os.path.join(results_dir, f'{base_filename}{extension}')
    
    # ファイルが存在する場合、インクリメントして一意な名前を生成
    counter = 1
    while os.path.exists(excel_file):
        excel_file = os.path.join(results_dir, f'{base_filename}({counter}){extension}')
        counter += 1
    
    return excel_file

# hostをキーにしてnameを取得する関数
def get_name_by_host(df, host):
    result = df[df['dest'] == host]['name']
    if not result.empty:
        return result.values[0]  # 一致するnameを返す
    else:
        return None  # 一致するhostが見つからない場合

# TODO: 出力先は各拠点のテストフォルダ結果内へ保存する
def write_results_to_excel(ping_results, trace_results, excel_file, df_test_eval):
    """
    pingとtraceの結果をExcelに書き出す関数。
    
    Parameters:
    - ping_results (list of dict): pingの結果 [{'host_a': True}, ...]
    - trace_results (list of dict): traceの結果 [{'host_a': False}, ...]
    - excel_file (str): 出力するExcelファイル名
    """
    # Excelに書き込むためのリストを作成
    data = []
    for i, (ping, trace) in enumerate(zip(ping_results, trace_results), start=1):
        host = list(ping.keys())[0]  # ホスト名取得
        # TODO: hostでdfを検索し該当するnameを取得する
        name = get_name_by_host(df_test_eval, host)
        ping_result = '○' if list(ping.values())[0] else '×'
        trace_result = '○' if list(trace.values())[0] else '×'
        data.append([i, name, host, ping_result, trace_result])

    # データをDataFrameに変換
    df = pd.DataFrame(data, columns=['num', 'name', 'dest', 'ping_result', 'trace_result'])

    # Excelファイルに書き出し
    df.to_excel(excel_file, index=False)

    print(f"Excelファイル '{excel_file}' に結果が書き込まれました。")
    
    return excel_file

def format_excel_file(excel_file):
    """
    Excelファイルのセルの幅を自動調整し、○と×のセルの色を変更する。

    Parameters:
    - excel_file (str): 処理するExcelファイルのパス
    """
    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active  # 最初のシートを取得

    # セルの幅を自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)  # 列のアルファベット取得
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # セルの幅に少し余裕を持たせる
        ws.column_dimensions[column_letter].width = adjusted_width

    # ○のセルは緑、×のセルは赤に設定
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 緑
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")      # 赤

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "○":
                cell.fill = green_fill
            elif cell.value == "×":
                cell.fill = red_fill

    # Excelファイルを保存
    wb.save(excel_file)
    print(f"'{excel_file}' のフォーマットが完了しました。")

def is_windows():
    """Windowsかどうかを判定"""
    return platform.system().lower() == 'windows'

def is_mac():
    """Macかどうかを判定"""
    return platform.system().lower() == 'darwin'

def get_os_language():
    """OSの言語設定を取得し、日本語版か英語版かを判定"""
    lang, _ = locale.getdefaultlocale()  # ロケール情報を取得
    
    if lang == 'ja_JP':
        return 'Japanese'
    elif lang.startswith('en'):
        return 'English'
    else:
        return 'Other'

def is_japanese_os():
    """OSが日本語版かどうかを判定"""
    return get_os_language() == 'Japanese'

def is_english_os():
    """OSが英語版かどうかを判定"""
    return get_os_language() == 'English'

class UnsupportedOSError(Exception):
    """サポートされていないOSのための例外"""
    pass

class UnsupportedLanguageError(Exception):
    """サポートされていない言語のための例外"""
    pass

def check_os_and_language():
    """
    OSと言語をチェックする関数。
    Windowsで日本語または英語がサポートされているか確認します。
    
    Returns:
        is_win (bool): Windowsかどうか
        is_ja (bool): 日本語OSかどうか
    """
    is_win = False
    is_ja = False

    if is_windows():
        is_win = True
        if is_japanese_os():
            is_ja = True
        elif is_english_os():
            pass
        else:
            raise UnsupportedLanguageError("サポートされていない言語です。")
    elif is_mac():
        pass
    else:
        raise UnsupportedOSError('サポートされていないOSです。')
    
    return is_win, is_ja
     
# テスト実行
def main():
    try:
        is_win, is_ja = check_os_and_language()
            
        # configファイルから設定値の読み込み
        ping_count, average_rtt, max_packet_loss, max_hop = load_config(config_path='../settings/setting.yml')

        # 拠点リストのファイルの読み込み
        df_kyoten = pd.read_csv('../settings/kyoten_list.csv')
        selected_kyoten_name, selected_kyoten_type = select_kyoten(df_kyoten)
        selected_test_type = select_test_type(selected_kyoten_type)
        
        # pandasで評価用CSVを読み込む
        df_test_eval = pd.read_csv(get_test_type_path(selected_kyoten_type, selected_test_type))

        # 複数のホストと期待されるpingの結果とtrace経路のリスト
        list_ping_eval, list_trace_eval = convert_to_list(df_test_eval)
        ping_results, trace_results, results_dir = asyncio.run(ping_trace_multiple_hosts(
            is_win = is_win,
            is_ja = is_ja,
            ping_count = ping_count,
            average_rtt = average_rtt,
            max_packet_loss = max_packet_loss,
            max_hop = max_hop,
            list_ping_eval = list_ping_eval,
            list_trace_eval = list_trace_eval,
            selected_kyoten_name = selected_kyoten_name,
            selected_test_type = selected_test_type))

        print(f'ping_results = {ping_results}')
        print(f'trace_results = {trace_results}')
        
        # 結果をExcelに書き込む
        excel_file = generate_unique_filename(results_dir, selected_kyoten_name, selected_test_type)
        writed_excel_file = write_results_to_excel(ping_results, trace_results, excel_file, df_test_eval)
        format_excel_file(writed_excel_file)
        
    except Exception as e:
        print(e)  # エラーメッセージを表示
        return  # 処理を終了

if __name__ == '__main__':
    main()